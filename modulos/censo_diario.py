import streamlit as st
import pandas as pd
import re
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font

# --- REGLAS DE NEGOCIO ---
ORDEN_TERAPIAS_EXCEL = ["UNIDAD CORONARIA", "UCIA", "TERAPIA POSQUIRURGICA", "U.C.I.N.", "U.T.I.P.", "UNIDAD DE QUEMADOS"]
MAPA_TERAPIAS = {"UNIDAD CORONARIA": "COORD_MODULARES", "U.C.I.N.": "COORD_PEDIATRIA", "U.T.I.P.": "COORD_PEDIATRIA", "TERAPIA POSQUIRURGICA": "COORD_MEDICINA", "UNIDAD DE QUEMADOS": "COORD_CIRUGIA", "UCIA": "COORD_MEDICINA"}
VINCULO_AUTO_INCLUSION = {"COORD_MEDICINA": ["UCIA", "TERAPIA POSQUIRURGICA"], "COORD_CIRUGIA": ["UNIDAD DE QUEMADOS"], "COORD_MODULARES": ["UNIDAD CORONARIA"], "COORD_PEDIATRIA": ["U.C.I.N.", "U.T.I.P."]}
COLORES_INTERFAZ = {"⚠️ UNIDADES DE TERAPIA ⚠️": "#C0392B", "COORD_PEDIATRIA": "#5DADE2", "COORD_MEDICINA": "#1B4F72", "COORD_GINECOLOGIA": "#F06292", "COORD_MODULARES": "#E67E22", "OTRAS_ESPECIALIDADES": "#2C3E50", "COORD_CIRUGIA": "#117864"}
CATALOGO = {
    "COORD_MEDICINA": ["DERMATO", "ENDOCRINO", "GERIAT", "INMUNO", "MEDICINA INTERNA", "REUMA", "UCIA", "TERAPIA INTERMEDIA", "CLINICA DEL DOLOR", "TPQX", "TERAPIA POSQUIRURGICA", "POSQUIRURGICA"],
    "COORD_CIRUGIA": ["CIRUGIA GENERAL", "CIR. GENERAL", "MAXILO", "RECONSTRUCTIVA", "PLASTICA", "GASTRO", "NEFROLOGIA", "OFTALMO", "ORTOPEDIA", "OTORRINO", "UROLOGIA", "TRASPLANTES", "QUEMADOS", "UNIDAD DE QUEMADOS"],
    "COORD_MODULARES": ["ANGIOLOGIA", "VASCULAR", "CARDIOLOGIA", "CARDIOVASCULAR", "TORAX", "NEUMO", "HEMATO", "NEUROCIRUGIA", "NEUROLOGIA", "ONCOLOGIA", "CORONARIA", "UNIDAD CORONARIA", "PSIQ", "PSIQUIATRIA"],
    "COORD_PEDIATRIA": ["PEDIATRI", "PEDIATRICA", "NEONATO", "NEONATOLOGIA", "CUNERO", "UTIP", "U.T.I.P", "UCIN", "U.C.I.N"],
    "COORD_GINECOLOGIA": ["GINECO", "OBSTETRICIA", "MATERNO", "REPRODUCCION", "BIOLOGIA DE LA REPRO"]
}

def obtener_especialidad_real(cama, esp_html):
    c = str(cama).strip().upper()
    esp_html_clean = esp_html.replace("ESPECIALIDAD:", "").replace("&NBSP;", "").strip().upper()
    if c.startswith("64"): return "UNIDAD CORONARIA"
    if c.startswith("55"): return "U.C.I.N."
    if c.startswith("45"): return "NEONATOLOGIA" 
    if c.startswith("56"): return "U.T.I.P."
    if c.startswith("85"): return "UNIDAD DE QUEMADOS"
    if c.startswith("73"): return "UCIA"
    if c.isdigit() and 7401 <= int(c) <= 7409: return "TERAPIA POSQUIRURGICA"
    return esp_html_clean

def sync_group(cat_name, servicios):
    master_val = st.session_state[f"master_{cat_name}"]
    for s in servicios: st.session_state[f"serv_{cat_name}_{s}"] = master_val

st.title("📋 Censo Epidemiológico Diario")

if 'archivo_compartido' not in st.session_state:
    st.info("👈 Por favor, sube el archivo HTML en el apartado de 'Configuración' de la izquierda.")
else:
    try:
        tablas = pd.read_html(st.session_state['archivo_compartido'])
        df_completo = max(tablas, key=len)
        col0_str = df_completo.iloc[:, 0].fillna("").astype(str).str.upper()
        
        pacs_detectados = []
        especialidades_encontradas = set()
        IGNORAR = ["PACIENTES", "TOTAL", "SUBTOTAL", "PÁGINA", "IMPRESIÓN", "1111"]
        
        esp_actual_temp = "SIN_ESPECIALIDAD"
        for i, val in enumerate(col0_str):
            if "ESPECIALIDAD:" in val: esp_actual_temp = val; continue
            fila = [str(x).strip() for x in df_completo.iloc[i].values]
            if any(x in fila[0] for x in IGNORAR): continue
            if len(fila[1]) >= 5 and any(char.isdigit() for char in fila[1]):
                esp_real = obtener_especialidad_real(fila[0], esp_actual_temp)
                especialidades_encontradas.add(esp_real)
                pacs_detectados.append({"CAMA": fila[0], "REG": fila[1], "PAC": fila[2], "SEXO": fila[3], "EDAD": "".join(re.findall(r'\d+', fila[4])), "DIAG": fila[6], "ING": fila[9], "esp_real": esp_real})

        st.subheader(f"📊 Pacientes Detectados: {len(pacs_detectados)}")

        # Organización en Buckets
        buckets = {}
        asignadas = set()
        terapias_list = sorted([e for e in especialidades_encontradas if e in ORDEN_TERAPIAS_EXCEL])
        if terapias_list:
            buckets["⚠️ UNIDADES DE TERAPIA ⚠️"] = terapias_list
            asignadas.update(terapias_list)
        
        for cat, kws in CATALOGO.items():
            found = sorted([e for e in especialidades_encontradas if e not in asignadas and any(kw in e for kw in kws)])
            if found:
                buckets[cat] = found
                asignadas.update(found)

        otras = sorted([e for e in especialidades_encontradas if e not in asignadas])
        if otras: buckets["OTRAS_ESPECIALIDADES"] = otras

        cols = st.columns(3)
        for idx, (cat_name, servicios) in enumerate(buckets.items()):
            with cols[idx % 3]:
                color = COLORES_INTERFAZ.get(cat_name, "#5D6D7E")
                st.markdown(f'<div style="background-color:{color}; padding:8px; border-radius:5px 5px 0px 0px; color:white; text-align:center;"><b>{cat_name.replace("COORD_", "")}</b></div>', unsafe_allow_html=True)
                with st.container(border=True):
                    st.checkbox(f"Seleccionar todo", key=f"master_{cat_name}", on_change=sync_group, args=(cat_name, servicios))
                    for s in servicios: st.checkbox(s, key=f"serv_{cat_name}_{s}")

        if st.button("🚀 GENERAR EXCEL", use_container_width=True, type="primary"):
            especialidades_finales = set()
            for c_name, servs in buckets.items():
                if st.session_state.get(f"master_{c_name}"):
                    if c_name in VINCULO_AUTO_INCLUSION:
                        for t in VINCULO_AUTO_INCLUSION[c_name]:
                            if t in especialidades_encontradas: especialidades_finales.add(t)
                for s in servs:
                    if st.session_state.get(f"serv_{c_name}_{s}"): especialidades_finales.add(s)

            if not especialidades_finales:
                st.warning("Selecciona al menos un servicio.")
            else:
                fecha_hoy = datetime.now()
                datos_excel = []
                for p in pacs_detectados:
                    if p["esp_real"] in especialidades_finales:
                        try:
                            f_ing = datetime.strptime(p["ING"], "%d/%m/%Y")
                            dias = (datetime(fecha_hoy.year, fecha_hoy.month, fecha_hoy.day) - datetime(f_ing.year, f_ing.month, f_ing.day)).days + 1
                        except: dias = "Rev."
                        datos_excel.append({"FECHA_REPORTE": fecha_hoy.strftime("%d/%m/%Y"), "ESPECIALIDAD": p["esp_real"], "CAMA": p["CAMA"], "REGISTRO": p["REG"], "PACIENTE": p["PAC"], "SEXO": p["SEXO"], "EDAD": p["EDAD"], "DIAGNOSTICO": p["DIAG"], "FECHA_INGRESO": p["ING"], "DIAS_ESTANCIA": dias})

                if datos_excel:
                    df_out = pd.DataFrame(datos_excel)
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_out.to_excel(writer, index=False, sheet_name='Epidemiologia')
                    
                    output.seek(0)
                    wb = load_workbook(output)
                    ws = wb.active
                    
                    # Definición de bordes delgados
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                    # Formateo dinámico y bordes
                    for col in ws.columns:
                        max_length = 0
                        column_letter = get_column_letter(col[0].column)
                        for cell in col:
                            cell.border = thin_border # Aplicar marcos a cada celda
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
                    
                    # Agregar Tabla Oficial
                    tab = Table(displayName="CensoTable", ref=ws.dimensions)
                    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                    ws.add_table(tab)
                    
                    final_io = BytesIO()
                    wb.save(final_io)
                    
                    # Mensaje solicitado
                    st.success("✅ Reporte de censo generado.") 
                    st.download_button(label="💾 DESCARGAR EXCEL", data=final_io.getvalue(), file_name=f"Censo_Epidemio_{fecha_hoy.strftime('%d%m%Y')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    except Exception as e:
        st.error(f"Error: {e}")
