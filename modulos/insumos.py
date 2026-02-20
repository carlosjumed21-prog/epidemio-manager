import streamlit as st
import pandas as pd
import re
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# --- FILTRO OFICIAL DE 11 ESPECIALIDADES ---
SERVICIOS_INSUMOS_FILTRO = [
    "HEMATOLOGIA", "HEMATOLOGIA PEDIATRICA", "ONCOLOGIA PEDIATRICA",
    "NEONATOLOGIA", "INFECTOLOGIA PEDIATRICA", "U.C.I.N.",
    "U.T.I.P.", "TERAPIA POSQUIRURGICA", "UNIDAD DE QUEMADOS",
    "ONCOLOGIA MEDICA", "UCIA"
]

def obtener_especialidad_real(cama, esp_html):
    c = str(cama).strip().upper()
    esp_html_clean = esp_html.replace("ESPECIALIDAD:", "").replace("&NBSP;", "").strip().upper()
    
    if c.startswith("55"): return "U.C.I.N."
    if c.startswith("45"): return "NEONATOLOGIA" 
    if c.startswith("56"): return "U.T.I.P."
    if c.startswith("85"): return "UNIDAD DE QUEMADOS"
    if c.startswith("73"): return "UCIA"
    if c.isdigit() and 7401 <= int(c) <= 7409: return "TERAPIA POSQUIRURGICA"
    return esp_html_clean

st.title("📦 Censo de Insumos")

if 'archivo_compartido' not in st.session_state:
    st.info("👈 Por favor, sube el archivo HTML en el apartado de 'Configuración' de la izquierda.")
else:
    try:
        tablas = pd.read_html(st.session_state['archivo_compartido'])
        df_completo = max(tablas, key=len)
        col0_str = df_completo.iloc[:, 0].fillna("").astype(str).str.upper()
        
        pacs_detectados = []
        IGNORAR = ["PACIENTES", "TOTAL", "SUBTOTAL", "PÁGINA", "IMPRESIÓN", "1111"]
        
        esp_actual_temp = "SIN_ESPECIALIDAD"
        for i, val in enumerate(col0_str):
            if "ESPECIALIDAD:" in val:
                esp_actual_temp = val
                continue
            
            fila = [str(x).strip() for x in df_completo.iloc[i].values]
            if any(x in fila[0] for x in IGNORAR): continue
            
            if len(fila[1]) >= 5 and any(char.isdigit() for char in fila[1]):
                esp_real = obtener_especialidad_real(fila[0], esp_actual_temp)
                
                if esp_real in SERVICIOS_INSUMOS_FILTRO:
                    pacs_detectados.append({
                        "CAMA": fila[0], "REG": fila[1], "PAC": fila[2], "SEXO": fila[3], 
                        "EDAD": "".join(re.findall(r'\d+', fila[4])), "ING": fila[9], 
                        "esp_real": esp_real
                    })

        if not pacs_detectados:
            st.warning("No se encontraron pacientes para las 11 especialidades de insumos.")
        else:
            servicios_encontrados = sorted(list(set([p["esp_real"] for p in pacs_detectados])))
            
            for serv in servicios_encontrados:
                with st.expander(f"🔍 Vista Previa: {serv}"):
                    df_p = pd.DataFrame([p for p in pacs_detectados if p["esp_real"] == serv])
                    df_p["TIPO DE PRECAUCIONES"] = df_p["esp_real"].apply(
                        lambda x: "ESTÁNDAR / PROTECTOR" if "ONCOLOGIA" in x or "QUEMADOS" in x else "ESTÁNDAR"
                    )
                    df_p["INSUMO"] = "JABÓN/SANITAS"
                    st.table(df_p[["CAMA", "REG", "PAC", "SEXO", "EDAD", "ING", "TIPO DE PRECAUCIONES", "INSUMO"]])

            if st.button("🚀 GENERAR EXCEL DE INSUMOS", use_container_width=True, type="primary"):
                hoy = datetime.now()
                venc = hoy + timedelta(days=7)
                f_hoy = hoy.strftime("%d/%m/%Y")
                f_venc = venc.strftime("%d/%m/%Y")
                
                output = BytesIO()
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    for serv in servicios_encontrados:
                        df_s = pd.DataFrame([p for p in pacs_detectados if p["esp_real"] == serv])
                        df_s["TIPO DE PRECAUCIONES"] = df_s["esp_real"].apply(
                            lambda x: "ESTÁNDAR / PROTECTOR" if "ONCOLOGIA" in x or "QUEMADOS" in x else "ESTÁNDAR"
                        )
                        df_s["INSUMO"] = "JABÓN/SANITAS"
                        
                        df_final = df_s[["CAMA", "REG", "PAC", "SEXO", "EDAD", "ING", "TIPO DE PRECAUCIONES", "INSUMO"]]
                        df_final.columns = ["CAMA", "REGISTRO", "PACIENTE", "SEXO", "EDAD", "FECHA DE INGRESO", "TIPO DE PRECAUCIONES", "INSUMO"]
                        
                        sheet_name = serv[:30].replace("/", "-")
                        df_final.to_excel(writer, index=False, sheet_name=sheet_name, startrow=1)
                        ws = writer.sheets[sheet_name]
                        
                        # --- ENCABEZADO SUPERIOR ---
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
                        cell_h = ws.cell(row=1, column=1, value=f"{serv} DEL {f_hoy} AL {f_venc} (PARA LOS 3 TURNOS Y FINES DE SEMANA)")
                        cell_h.alignment = Alignment(horizontal="center", vertical="center")
                        cell_h.font = Font(bold=True)

                        # --- CUERPO: BORDES Y AUTOAJUSTE ---
                        lr = ws.max_row
                        for row in ws.iter_rows(min_row=2, max_row=lr, min_col=1, max_col=8):
                            for cell in row:
                                cell.border = thin_border
                                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

                        for i, col_name in enumerate(df_final.columns):
                            L = get_column_letter(i + 1)
                            max_len = len(col_name)
                            for cell in ws[L]:
                                if cell.row > 1 and cell.value:
                                    max_len = max(max_len, len(str(cell.value)))
                            ws.column_dimensions[L].width = min(max_len + 4, 40)

                        # --- PIE DE PÁGINA: NOM-045 ---
                        ws.merge_cells(start_row=lr + 1, start_column=1, end_row=lr + 1, end_column=8)
                        leyenda = "Comentario: de acuerdo con la Norma Oficial Mexicana NOM-045-SSA2-2005, Para la vigilancia epidemiológica, prevención y control de las infecciones nosocomiales. NINGUN RECIPIENTE QUE CONTENGA EL INSUMO DEVERÁ SER RELLENADO O REUTILIZADO."
                        cell_f = ws.cell(row=lr + 1, column=1, value=leyenda)
                        cell_f.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell_f.font = Font(size=9, italic=True)
                        ws.row_dimensions[lr + 1].height = 55 
                        
                        # --- FIRMA: AUTORIZÓ (JUSTO DEBAJO) ---
                        # Usamos lr + 2 para eliminar la fila de espacio
                        ws.merge_cells(start_row=lr + 2, start_column=1, end_row=lr + 2, end_column=8)
                        cell_auth = ws.cell(row=lr + 2, column=1, value="AUTORIZÓ: DRA. BRENDA CASTILLO MATUS")
                        cell_auth.alignment = Alignment(horizontal="center", vertical="center")
                        cell_auth.font = Font(bold=True)

                st.success("✅ Reporte de insumos generado.")
                st.download_button(
                    label="💾 DESCARGAR REPORTE DE INSUMOS", 
                    data=output.getvalue(), 
                    file_name=f"Insumos_Epidemio_{hoy.strftime('%d%m%Y')}.xlsx",
                    use_container_width=True
                )
    except Exception as e:
        st.error(f"Error: {e}")
